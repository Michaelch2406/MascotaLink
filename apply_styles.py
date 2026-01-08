import openpyxl
from copy import copy

file_path = "Matriz de trazabilidad.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit(1)

# Usaremos la fila 5 (REQ-004) como plantilla de estilo
source_row_idx = 5
# Las nuevas filas van de la 6 a la 15
start_target_row = 6
end_target_row = 15

print(f"Applying styles from row {source_row_idx} to rows {start_target_row}-{end_target_row}...")

for row in range(start_target_row, end_target_row + 1):
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row_idx, column=col)
        target_cell = ws.cell(row=row, column=col)

        if source_cell.has_style:
            # OpenPyXL requiere copiar los objetos de estilo
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

wb.save(file_path)
print("Styles applied successfully.")
